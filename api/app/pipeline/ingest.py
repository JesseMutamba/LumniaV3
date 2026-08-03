"""Layer 01 — ingestion. Files as the client already keeps them.

Deliberately dumb. This module opens a workbook, inventories it, and hands
back a grid plus a Source record. It does not try to understand the sheet;
that is layer 02's job and layer 02 is not built yet.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..schema import Source


@dataclass
class Sheet:
    name: str
    grid: list[list]  # row-major, 1-indexed rows map to grid[r-1]

    def cell(self, row: int, col: int):
        """1-indexed, spreadsheet convention. Returns None out of range."""
        try:
            return self.grid[row - 1][col - 1]
        except IndexError:
            return None

    def col_values(self, col: int, r0: int, r1: int) -> list:
        return [self.cell(r, col) for r in range(r0, r1 + 1)]

    @property
    def rows(self) -> int:
        return len(self.grid)


@dataclass
class Workbook:
    path: Path
    sheets: dict[str, Sheet]
    source: Source

    def __getitem__(self, name: str) -> Sheet:
        return self.sheets[name]


def a1(row: int, col: int) -> str:
    """(4, 3) -> 'C4'. Cheap, and it keeps provenance strings honest."""
    s = ""
    while col:
        col, rem = divmod(col - 1, 26)
        s = chr(65 + rem) + s
    return f"{s}{row}"


def a1_range(r0: int, c0: int, r1: int, c1: int) -> str:
    return f"{a1(r0, c0)}:{a1(r1, c1)}" if (r0, c0) != (r1, c1) else a1(r0, c0)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def read_workbook(path: str | Path, idx: int = 0) -> Workbook:
    """Read every sheet into a plain grid of evaluated values.

    data_only=True gives us the cached formula results Excel last wrote. If a
    workbook has never been opened by Excel those come back None — we surface
    that as an empty sheet rather than guessing, because a guessed number is
    exactly the thing this platform exists to not do.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, Sheet] = {}
    total_rows = 0

    for ws in wb.worksheets:
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # trim trailing blank rows — openpyxl pads generously
        while grid and not any(c is not None for c in grid[-1]):
            grid.pop()
        sheets[ws.title] = Sheet(ws.title, grid)
        total_rows += len(grid)

    wb.close()
    return Workbook(
        path=path,
        sheets=sheets,
        source=Source(
            idx=idx,
            filename=path.name,
            sha256=sha256(path),
            sheets=len(sheets),
            rows_read=total_rows,
        ),
    )
